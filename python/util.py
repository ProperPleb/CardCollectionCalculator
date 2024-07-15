import constants as C
import logging
import re
import os.path

from python.request_dto import *
from openpyxl import *
from card import Card

logger = logging.getLogger(__name__)


class GridUtil:

    @staticmethod
    def instantiateGridDTO() -> GridDTO:
        grid = GridDTO()
        grid.algorithm = C.DEFAULT_ALGO
        grid.from_ = C.DEFAULT_FROM
        grid.size = C.DEFAULT_GRID_SIZE
        grid.filters = GridUtil.instantiateFilter("GRID")
        grid.listingSearch = GridUtil.instantiateListingSearch()
        grid.context = GridUtil.instantiateContext()
        grid.settings = GridUtil.instantiateSettings()
        grid.sort = GridUtil.instantiateSort()
        return grid

    @staticmethod
    def instantiateContext() -> Context:
        context = Context()
        context.cart = {}
        context.shippingCountry = C.COUNTRY_CODE_US
        user_profile = UserProfile()
        user_profile.priceAffinity = C.DEFAULT_PRICE_AFFINITY
        context.userProfile = user_profile
        return context

    @staticmethod
    def instantiateFilter(level: str) -> Filters:
        if level == "GRID":
            return GridUtil.instantiateGridFilter(level)
        elif level == "LISTING_SEARCH":
            return GridUtil.instantiateLSFilter(level)
        return Filters()

    @staticmethod
    def instantiateGridFilter(level: str) -> Filters:
        filters = Filters()
        term = GridUtil.instantiateTerm(level)
        filters.term = term
        filters.range = Range()
        filters.match_ = {}
        return filters

    @staticmethod
    def instantiateTerm(level: str) -> Term:
        if level == "GRID":
            return GridUtil.instantiateGridTerm()
        elif level == "LISTING_SEARCH":
            return GridUtil.instantiateLSTerm()
        return Term()

    @staticmethod
    def instantiateGridTerm() -> Term:
        term = Term()
        term.productLineName = [C.YUGIOH_LOWER]
        return term

    @staticmethod
    def instantiateListingSearch() -> ListingSearch:
        listing_search = ListingSearch()
        context = Context()
        context.cart = {}
        listing_search.context = context
        listing_search.filters = GridUtil.instantiateFilter("LISTING_SEARCH")
        return listing_search

    @staticmethod
    def instantiateLSFilter(level: str) -> Filters:
        filters = Filters()
        filters.term = GridUtil.instantiateTerm(level)
        range_ = Range()
        quantity = Quantity()
        quantity.gte = C.DEFAULT_GTE
        range_.quantity = quantity
        filters.range_ = range_
        exclude = Exclude()
        exclude.channelExclusion = C.DEFAULT_CHANNEL_EXCLUSION
        filters.exclude = exclude
        return filters

    @staticmethod
    def instantiateLSTerm() -> Term:
        term = Term()
        term.sellerStatus = C.SELLER_STATUS_LIVE
        term.channelId = C.DEFAULT_CHANNEL_ID
        term.language = [C.LANGUAGE_ENGLISH]
        term.condition = [C.CONDITION_NM]
        term.listingType = [C.DEFAULT_LISTING_TYPE]
        return term

    @staticmethod
    def instantiateSettings() -> Settings:
        settings = Settings()
        settings.useFuzzySearch = C.DEFAULT_FUZZY_SEARCH
        settings.didYouMean = {}
        return settings

    @staticmethod
    def instantiateSort() -> Sort:
        sort = Sort()
        sort.field = C.DEFAULT_GRID_SORT_FIELD
        sort.order = C.DEFAULT_SORT_ORDER
        return sort


class DetailViewUtil:

    @staticmethod
    def instantiateDetailViewDTO() -> DetailViewDTO:
        detail_view = DetailViewDTO()
        detail_view.context = DetailViewUtil.instantiateContext()
        detail_view.filters = DetailViewUtil.instantiateFilters()
        detail_view.from_ = C.DEFAULT_FROM
        detail_view.size = C.DEFAULT_DETAIL_VIEW_SIZE
        detail_view.sort = DetailViewUtil.instantiateSort()
        return detail_view

    @staticmethod
    def instantiateContext() -> Context:
        context = Context()
        context.cart = {}
        shippingCountry: C.COUNTRY_CODE_US
        return context

    @staticmethod
    def instantiateFilters() -> Filters:
        filters = Filters()
        exclude = Exclude()
        exclude.channelExclusion = C.DEFAULT_CHANNEL_EXCLUSION
        filters.exclude = exclude
        range_ = Range()
        quantity = Quantity()
        quantity.gte = C.DEFAULT_GTE
        range_.quantity = quantity
        filters.range_ = range_
        term = Term()
        term.channelId = C.DEFAULT_CHANNEL_ID
        term.condition = [C.CONDITION_NM]
        term.language = [C.LANGUAGE_ENGLISH]
        term.listingType = [C.DEFAULT_LISTING_TYPE]
        term.printing = [C.PRINTING_1ST_ED]
        term.sellerStatus = C.SELLER_STATUS_LIVE
        filters.term = term
        return filters

    @staticmethod
    def instantiateSort() -> Sort:
        sort = Sort()
        sort.field = C.DEFAULT_DETAIL_VIEW_SORT_FIELD
        sort.order = C.DEFAULT_SORT_ORDER
        return sort


class JSONUtil:

    @staticmethod
    def overrideVariableNames(string: str) -> str:
        string = string.replace("range_", "range")
        string = string.replace("match_", "match")
        string = string.replace("from_", "from")
        return string


class ExcelUtil:

    ORIGINAL_HEADER_NAMES: list

    @staticmethod
    def excelMapper(file_path: str = None, file_name: str = None) -> list:
        ExcelUtil.ORIGINAL_HEADER_NAMES = []
        if file_name is None:
            file_name = C.DEFAULT_EXCEL_FILE_NAME
        if file_path is None:
            file_path = C.DEFAULT_EXCEL_FILE_PATH

        wb: Workbook = load_workbook(file_path + file_name)
        sheet = wb.worksheets[0]
        headers: list = []
        card_list: list = []
        current_cell = sheet['A1']

        while current_cell.value is not None:
            ExcelUtil.ORIGINAL_HEADER_NAMES.append(str(current_cell.value))
            headers.append(str(current_cell.value).lower().strip().replace(" ", "_", -1))
            current_cell = current_cell.offset(0, 1)

        first_cell = sheet['A2']
        current_cell = first_cell
        is_empty: bool = False

        while not is_empty:
            is_empty = True
            card_dict = {}
            for i in range(len(headers)):
                if current_cell.value is not None:
                    is_empty = False
                card_dict.update({headers[i]: current_cell.value})
                current_cell = current_cell.offset(0, 1)
            if not is_empty and ExcelUtil.sanitize(card_dict):
                card_list.append(Card(**card_dict))
            current_cell = first_cell.offset(1, 0)
            first_cell = current_cell
        wb.close()
        return card_list

    @staticmethod
    def saveToExcel(card_list: list, file_path: str = None, file_name: str = None) -> bool:
        if file_name is None:
            file_name = C.DEFAULT_EXCEL_FILE_NAME
        if file_path is None:
            file_path = C.DEFAULT_EXCEL_FILE_PATH

        wb: Workbook
        if os.path.exists(file_path + file_name):
            wb = load_workbook(file_path + file_name)
        else:
            wb = Workbook()
            wb.create_sheet('Sheet1', 0)

        sheet = wb.worksheets[0]
        first_cell = sheet['A1']
        current_cell = first_cell
        headers: list = []

        for header in ExcelUtil.ORIGINAL_HEADER_NAMES:
            header = str(header)
            current_cell.value = header
            headers.append(header.lower().strip().replace(" ", "_", -1))
            current_cell = current_cell.offset(0, 1)

        first_cell = sheet['A2']
        current_cell = first_cell
        card_list = ExcelUtil.reMap(card_list)
        for card in card_list:
            card_dict = card.__dict__
            for header in headers:
                current_cell.value = card_dict[header]
                current_cell = current_cell.offset(0, 1)
            first_cell = first_cell.offset(1, 0)
            current_cell = first_cell

        wb.save(file_path + file_name)
        wb.close()
        return True

    @staticmethod
    def sanitize(entry: dict) -> bool:
        unit_price = entry[C.EXCEL_HEADER_UNIT_PRICE]
        recalc = entry[C.EXCEL_HEADER_RECALC]
        if unit_price is not None and unit_price > 0 and recalc is None:
            return False

        card_name: str = entry[C.EXCEL_HEADER_CARD_NAME].strip()
        if not card_name:
            return False

        if recalc is not None:
            entry[C.EXCEL_HEADER_RECALC]: bool = True
        else:
            entry[C.EXCEL_HEADER_RECALC]: bool = False

        quantity = entry[C.EXCEL_HEADER_QUANTITY]
        if quantity is None or not isinstance(quantity, int) or quantity < 1:
            entry[C.EXCEL_HEADER_QUANTITY]: int = 0

        rarity = entry[C.EXCEL_HEADER_RARITY]
        if rarity is not None and isinstance(rarity, str) and C.RARITY[rarity] is not None:
            entry[C.EXCEL_HEADER_RARITY]: str = str(C.RARITY[rarity])
        else:
            entry[C.EXCEL_HEADER_RARITY]: str = None

        card_name = entry[C.EXCEL_HEADER_CARD_NAME].strip()
        if card_name is not None and isinstance(card_name, str):
            card_name = re.sub(r'[^A-Za-z0-9 ]+', '', card_name)
            card_name = re.sub(r' +', ' ', card_name)
            entry[C.EXCEL_HEADER_CARD_NAME]: str = card_name
        else:
            entry[C.EXCEL_HEADER_CARD_NAME]: str = None

        if unit_price is None or (not isinstance(unit_price, int) and not isinstance(unit_price, float)) or unit_price <= 0:
            entry[C.EXCEL_HEADER_UNIT_PRICE]: float = float(0.0)
        else:
            entry[C.EXCEL_HEADER_UNIT_PRICE]: float = float(unit_price)

        total_price = entry[C.EXCEL_HEADER_TOTAL_PRICE]
        if total_price is None or (not isinstance(total_price, int) and not isinstance(total_price, float)) or total_price <= 0:
            entry[C.EXCEL_HEADER_TOTAL_PRICE]: float = float(0.0)
        else:
            entry[C.EXCEL_HEADER_TOTAL_PRICE]: float = float(total_price)

        edition = entry[C.EXCEL_HEADER_EDITION]
        if edition is not None and isinstance(edition, str) and C.PRINTING[edition] is not None:
            entry[C.EXCEL_HEADER_EDITION]: str = C.PRINTING[edition]
        else:
            entry[C.EXCEL_HEADER_EDITION]: str = None

        condition = entry[C.EXCEL_HEADER_CONDITION]
        if condition is not None and isinstance(condition, str) and C.CONDITION[condition] is not None:
            entry[C.EXCEL_HEADER_CONDITION]: str = C.CONDITION[condition]
        else:
            entry[C.EXCEL_HEADER_CONDITION]: str = None

        return True

    @staticmethod
    def reMap(card_list: list):
        rarity = {v: k for k, v in C.RARITY.items()}
        edition = {v: k for k, v in C.PRINTING.items()}
        condition = {v: k for k, v in C.CONDITION.items()}

        for card in card_list:
            card.recalc = None
            if card.rarity is not None:
                card.rarity = rarity[card.rarity]
            if card.edition is not None:
                card.edition = edition[card.edition]
            if card.condition is not None:
                card.condition = condition[card.condition]

        return card_list
